#!/usr/bin/python
# -*- coding: utf-8 -*-

'''
Desc: create a excel, read data from it, and save them to the datatbase, remain to refactoring.
Author: Kris Peng
Copyright (c) 2019 - Kris Peng <kris.dacpc@gmail.com>
'''
from openpyxl import load_workbook
from pandas import DataFrame
from sqlalchemy import Column, String, Integer, create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.declarative import declarative_base

# 创建对象的基类:
Base = declarative_base()

class CASE(Base):
    __tablename__ = 'case_info'
    id = Column(Integer, primary_key=True)
    case_id = Column(String(64))
    name = Column(String(64))
    phone_number = Column(String(64))
    title = Column(String(1024))
    content = Column(String(4096))
    remark = Column(String(4096))
    status = Column(String(64))
    ctype = Column(String(64))
    address = Column(String(1024))
    case_status = Column(String(64))

    def __init__(self, case_id, name, phone_number, title, content, remark, status, ctype, address, case_status):
        self.case_id = case_id
        self.name = name
        self.phone_number = phone_number
        self.title = title
        self.content = content
        self.remark = remark
        self.status = status
        self.ctype = ctype
        self.address = address
        self.case_status = case_status

dest_filename = 'BL_C_INFO.xlsx'
wb = load_workbook(dest_filename)
ws = wb.get_sheet_by_name('BL_C_INFO(2)')
list = []
list_count = 0
for columns in ws.iter_rows(row_offset = 1):
    cur_list = []
    list_count = list_count + 1
    print("s"+str(list_count))
    for column in columns:
        cur_list.append(str(column.value))
    list.append(cur_list)
engine = create_engine('mysql+pymysql://root:pengcheng00@localhost:3306/taicang?charset=utf8mb4')
DBSession = sessionmaker(bind=engine)
session = DBSession()
count = 0
for row in list:
    count = count + 1
    print(count)
    new_case = CASE(case_id = row[0], 
                    name = row[1],
                    phone_number = row[2],
                    title = row[3],
                    content = row[4],
                    remark = row[5],
                    status = row[6],
                    ctype = row[7],
                    address = row[8],
                    case_status = row[9])
    session.add(new_case)
    session.commit()
session.close()