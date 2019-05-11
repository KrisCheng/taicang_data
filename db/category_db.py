#!/usr/bin/python
# -*- coding: utf-8 -*-

'''
Desc: 事件类型入库。
Author: Kris Peng
Copyright (c) 2019 - Kris Peng <kris.dacpc@gmail.com>
'''

from openpyxl import load_workbook
from pandas import DataFrame
from sqlalchemy import Column, String, Integer, create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.declarative import declarative_base
from config.config import *
from model.casecategory import CaseCategory

def main_task():
    dest_filename = 'data/SJFL.xlsx'
    wb = load_workbook(dest_filename)
    ws = wb.get_sheet_by_name('SJFL')
    list = []
    list_count = 0
    for columns in ws.iter_rows(row_offset = 1):
        cur_list = []
        list_count = list_count + 1
        print("s "+str(list_count))
        for column in columns:
            cur_list.append(str(column.value))
        list.append(cur_list)
    engine = create_engine(MYSQL_DATABASE_URI)
    DBSession = sessionmaker(bind=engine)
    session = DBSession()
    count = 0
    for row in list:
        count = count + 1
        print(count)
        new_caseclass = CaseCategory(category = row[0], 
                        categorycode = row[1],
                        remark = row[2])
        session.add(new_caseclass)
        session.commit()
    session.close()