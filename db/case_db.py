#!/usr/bin/python
# -*- coding: utf-8 -*-

'''
Desc: 数据入库。
Author: Kris Peng
Copyright (c) 2019 - Kris Peng <kris.dacpc@gmail.com>
'''
from openpyxl import load_workbook
from pandas import DataFrame
from sqlalchemy import Column, String, Integer, create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.declarative import declarative_base
from config.config import *

def main_task():
    dest_filename = 'BL_C_INFO.xlsx'
    wb = load_workbook(dest_filename)
    ws = wb.get_sheet_by_name('BL_C_INFO(2)')
    list = []
    list_count = 0
    for columns in ws.iter_rows(row_offset = 1):
        cur_list = []
        list_count = list_count + 1
        print("s "+str(list_count))
        for column in columns:
            cur_list.append(str(column.value))
        list.append(cur_list)
    engine = create_engine(MYSQL_DATABASE_URI)
    DBSession = sessionmaker(bind=engine)
    session = DBSession()
    count = 0
    for row in list:
        count = count + 1
        print(count)
        new_case = Case(case_id = row[0], 
                        name = row[1],
                        phone_number = row[2],
                        title = row[3],
                        content = row[4],
                        remark = row[5],
                        status = row[6],
                        ctype = row[7],
                        address = row[8],
                        case_status = row[9])
        session.add(new_case)
        session.commit()
    session.close()